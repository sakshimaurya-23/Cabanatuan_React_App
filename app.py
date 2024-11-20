from flask import Flask, request, jsonify, send_file, Response
import openpyxl
from flask_cors import CORS
import os
import pandas as pd
import matplotlib.pyplot as plt
import sqlite3
import re 
import concurrent.futures
from openpyxl import load_workbook
from dotenv import load_dotenv
from langchain_ibm import WatsonxLLM
import plotly.express as px
import os



desc_file_path = "description1.csv"
database_name = "main.db"

def create_SQL_database(file_path, database_name):
    file = pd.ExcelFile(file_path)
    conn = sqlite3.connect(database_name)
    c = conn.cursor()
    for sheet_name in file.sheet_names:
        df = pd.read_excel(file, sheet_name=sheet_name)
        df.columns = df.columns.astype(str)
        df.columns = df.columns.str.replace(' ', '_')
        table_name = sheet_name.replace(" ", "_")
        df.to_sql(table_name, conn, if_exists='replace', index=False)
        print(f"Uploaded sheet '{sheet_name}' as table '{table_name}'")
    conn.commit()
    c.close()
    print("All sheets uploaded")

def get_table_description(file_path):
    df = pd.read_csv(file_path)
    print(df)
    table_desc = ""
    for index, row in df.iterrows():
        table_desc += "\nTable Name : " + row[0]
        table_desc += "\nTable Description : " + row[1]
    return table_desc


def get_table_content(file_path):
    wb = load_workbook(file_path)
    structured_string = ""
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        table_name = sheet_name.replace(' ', '_')
        structured_string += f"Table Name: {table_name}\n"
        max_row = sheet.max_row
        max_col = sheet.max_column
        headers = [sheet.cell(row=1, column=col).value for col in range(1, max_col + 1)]
        structured_string += "Columns:\n" + ", ".join(str(header.replace(' ','_')) for header in headers if header) + "\n"
        for row_num in range(2, 7):
            structured_string += f"\nRow {row_num - 1} data:\n"
            for col in range(1, max_col + 1):
                cell_value = sheet.cell(row=row_num, column=col).value
                structured_string += f"{headers[col-1].replace(' ','_')}: {cell_value}\n" if headers[col-1] else f"Column {col}: {cell_value}\n"
        structured_string += "\n\n"
    return structured_string


def create_agents():
    # project_id = project_id
    parameters_1 = {
        "decoding_method": 'greedy',
        "max_new_tokens": 4000,
        "min_new_tokens": 5,
        "temperature": 0
    }
    parameters_2 = {
        "decoding_method": 'greedy',
        "max_new_tokens": 4000,
        "min_new_tokens": 5,
        "temperature": 0,
        "repetition_penalty":1.13
    }
    sql_model_id = 'mistralai/mistral-large' #meta-llama/llama-3-70b-instruct
    reviewer_model_id = 'mistralai/mistral-large' #mistralai/mistral-large
    sql_llm = WatsonxLLM(
        model_id = sql_model_id,
        project_id =os.getenv("project_id"),
        apikey=os.getenv("WATSONX_APIKEY"),
        url=os.getenv("WATSONX_URL"),
        params = parameters_1
    )
    reviewer_llm = WatsonxLLM(
        model_id = reviewer_model_id,
        project_id =os.getenv("project_id"),
        apikey=os.getenv("WATSONX_APIKEY"),
        url=os.getenv("WATSONX_URL"),
        params=parameters_2
    )
    return sql_llm, reviewer_llm

def execute_SQL_query(database_name, query):
    """Executes the SQL query and returns the results"""
    conn = sqlite3.connect('main.db')
    result = pd.read_sql_query(query, conn).to_string()
    conn.close()
    return result

def generate_SQL_query(table_description, table_content, user_input, sql_llm):

    print("""Generating SQL query""")
    print("User Input: ",user_input)

    prefix_sql = f"""As an expert SQL query writer specializing in SQLite, your goal is to craft efficient, optimized SQL queries tailored to user needs. Your queries should seamlessly interact with SQLite databases, making use of appropriate type casting to float when calculating percentages. Avoid using multiple JOIN statements directly in the main query. Instead, use Common Table Expressions (CTEs) or subqueries to simplify complex join logic and enhance readability. Note that SQLite does not support RIGHT JOIN or OUTER JOIN.

Example structure for queries involving aggregated metrics:
    WITH EntityAggregates AS (
        SELECT
            EntityID,
            SUM(Value) AS TotalValue
        FROM
            EntityTable
        GROUP BY
            EntityID
    ),
    TargetAggregates AS (
    SELECT
        EntityID,
        SUM(Target) AS TotalTarget
    FROM
        TargetTable
    WHERE
        TimePeriod = 'SpecificPeriod'
    GROUP BY
        EntityID
    )
    SELECT
        E.EntityName,
        COALESCE(EA.TotalValue, 0) AS TotalValue,
        COALESCE(TA.TotalTarget, 0) AS TotalTarget,
        ROUND((COALESCE(EA.TotalValue, 0) / CAST(COALESCE(TA.TotalTarget, 0) AS FLOAT)) * 100, 2) AS PercentageAchievement
    FROM
        EntityDetails E
    LEFT JOIN
        EntityAggregates EA ON E.EntityID = EA.EntityID
    LEFT JOIN
        TargetAggregates TA ON E.EntityID = TA.EntityID
    ORDER BY
        PercentageAchievement DESC;
    Ensure that you use the column names exactly as provided in the information. Output the SQL query as a string.
    YOUR RESPONSE MUST CONTAIN ONLY THE SQL QUERY.

    ONLY ONE SQL QUERY as a string. 
    
    Question : {user_input}
    Table Description : {table_description}
    Table Columns and Sample rows : {table_content}
    SQL Query : """

    sql_response = sql_llm.invoke(prefix_sql).strip()
    pattern = re.compile(r'SQL_QUERY\s*=\s*"""\s*(.*?)\s*"""', re.DOTALL)

    match = pattern.search(sql_response)

    if match:
        sql_response = match.group(1)
    sql_response = sql_response.replace('```sql', '').replace('```', '').strip()
    print(sql_response)
    return sql_response


def validate_SQL_query(table_description, table_content, user_input, query, reviewer_llm):
    print("In Validation SQl Query Function")
    
    prefix_validation = f""" As a professional SQL specialist, you have been assisting users with refining their SQL queries for various projects and tasks. Your ability to identify and correct syntax errors, optimize query performance, and ensure logical accuracy has earned you a stellar reputation in the industry.
    Your task today is to review and rewrite a SQL query for a user. The user will provide you with their initial query and any specific requirements they have. It is crucial that you pay attention to details, adhere to best practices, and enhance the query's efficiency while maintaining its intended functionality.
        
    Important Points to Consider:
    1. Verify the SQL query is based solely on the provided information.
    2. Do not make assumptions beyond the provided details, especially regarding IDs.
    3. Sample rows are for context only and should not be used as actual data.
    4. If the query involves finding an ID based on a name, first extract the ID from the appropriate table.
    5. Optimize the query and ensure proper use of indexes.
    6. Ensure compliance with SQLite rules, especially when working with date functions (use `strftime` instead of `MONTH()`).
    7. Type cast to float when calculating percentages.
    8. Ensure all columns used belong to their respective tables.
    9. Avoid using RIGHT JOIN or OUTER JOIN as they are not supported by SQLite.
    10. For complex join logic, consider using Common Table Expressions (CTEs) to break down the query.
    11. If the query involves more than two JOIN statements, provide a different approach using subqueries, CTEs, or breaking down the query into smaller parts.
     Example for complex joins:
        Original:
        SELECT A.Col1, B.Col2, C.Col3 
        FROM TableA A 
        JOIN TableB B ON A.ID = B.A_ID 
        JOIN TableC C ON B.ID = C.B_ID 
        WHERE A.Condition = 'Value';

        Revised using CTEs:
        WITH CTE1 AS (
            SELECT A.ID, A.Col1 
            FROM TableA A 
            WHERE A.Condition = 'Value'
        ),
        CTE2 AS (
            SELECT B.A_ID, B.Col2 
            FROM TableB B 
            JOIN CTE1 ON CTE1.ID = B.A_ID
        )
        SELECT CTE1.Col1, CTE2.Col2, C.Col3 
        FROM TableC C 
        JOIN CTE2 ON CTE2.B_ID = C.B_ID;
The query is correct only if it satisfies all these conditions. If it is not correct, provide the corrected SQL query.

Output only the SQL query. Do not provide any explanation.

    
    Only if the query satisfies all these conditions it is correct.
    Is the SQL query correct? If not, please provide the correct SQL query.
    Output the SQL query. Even the Given Query is Correct output the same query.
    Do not provide any explanation. Only the query.
    ONLY OUTPUT THE SINGLE CORRECT QUERY.

    Question: {user_input}
    Table Description: {table_description}
    Table Columns and Sample Rows: {table_content}
    SQL Query: {query}

    Corrected SQL Query: """
    reviewer_response = reviewer_llm.invoke(prefix_validation)
    validated_query = re.sub(r'"""', '', reviewer_response)
    validated_query = validated_query.replace('```sql', '').replace('```', '').strip()
    validated_query = validated_query.split(';')[0] + ';'
    print(validated_query)
    return validated_query


def create_visualization(database_name, user_input, query, graph_llm, html_file):

    prompt = f"""You're a seasoned data visualization expert with a keen eye for creating visually appealing graphs that effectively convey insights from complex datasets. Your expertise lies in leveraging Python libraries to generate interactive and informative graphs based on dynamic data inputs. Your task is to assist me in generating a graph or table using Python code that visualizes the results of a SQL query. Here are the details I will provide to you for creating the graph or table:

    Input Question: {user_input}
    SQL Query: {query}

    If the results on executing the query less than 2 rows, then you must plot a table of the results. Else, you must plot a graph of the results.
    If the results on executing the query has 2 columns and 2-5 rows, then you must use a pie chart.
    The graph or table should be clear, concise, and visually appealing, providing meaningful insights into the data extracted from the SQL query. Incorporate appropriate title, labels, titles, and formatting to enhance the overall presentation of the graph. 
    If needed, you can refer to past examples of similar tasks to ensure that the result is not only functional but also aesthetically pleasing and easy to interpret.

    The tables are stored in the {database_name} file in the same parent folder. Save the plot or table as {html_file}. The query must be stored in a variable. The code should be ready for execution with no placeholders requiring manual input.
    Use go.Table from plotly.graph_objects to generate the table visuals.

    Use Plotly for visualization. Output only the code as a string without explanations. Ensure proper indentation for code blocks under if else statements. Do not use unnecessary indents or backticks.
    """

    code_response = graph_llm.invoke(prompt).strip()

    lines = code_response.split('\n')
    lines = [line for line in lines[1:-1]]
    code_response = '\n'.join(lines)
    pattern = r'```python'
    cleaned_code = re.sub(pattern, '', code_response, flags=re.IGNORECASE)
    print(cleaned_code)
    try:
        exec(cleaned_code,locals())
    except Exception as e:
        print("Error in generating chart. The error was ", e)
    return cleaned_code


def generate_response(reviewer_llm, user_input, result):
    prompt = f"""
        Given the user question, table description and the results of runnning an SQL query, 
        generate an appropriate natural language response based on the results and user question.
        Return only the response which is comprehendable by the user. Do not provide unnecessary details such as table and column description.
        Please don't generate python code as answer.
        Question : {user_input}
        Results: {result}
        Response:
    """
    NL_response = reviewer_llm.invoke(prompt).strip()
    pattern = r'^Response:\s*(.*)'
    match = re.search(pattern, NL_response)
    if match:
        extracted_text = match.group(1)
        NL_response = extracted_text
    return NL_response


def generate_questions(reviewer_llm, table_description, table_content):

    prompt = f"""
    Generate four concise, descriptive questions that focus on analyzing sales data based on the description and sample data provided below.
    
    Table Description: {table_description}
    Table Content: {table_content}

    
    Question Style: 
    Do not ask the highest or lowest value.
    Include the word 'across' or 'each' or 'various'.

    Craft four questions that align with the rules specified and tailored to the specific regional focus.
    """

    questions = reviewer_llm.invoke(prompt)
    return questions


def source_attribution(question,query,reviewer_llm,nlp_ans):
    prompt='''When provided with an SQL query, extract and return the names of the tables and the columns used from each table. Present the information clearly in sentence format, without any additional symbols or hashtags.'''
    global tables_used
    tables_used=reviewer_llm.invoke(prompt+query)
    file_path = 'src/metadata.txt'
    # file_path = '/Users/sakshimaurya/Desktop/rohan bhandari talk to my data /Talk-To-My-Data/src/metadata.txt'
    with open(file_path, 'w') as file:
        file.write("Question:\n")
        file.write(question + '\n\n')
        file.write("SQL Query:\n")
        file.write(query + '\n\n')
        file.write("Table_Info:")
        file.write(tables_used + '\n')
        file.write("\nResponse:\n")
        file.write(nlp_ans)
    print(f"Content successfully written to {file_path}")

def execute_SQL_query(database_name, query):
    """Executes the SQL query and returns the results"""
    conn = sqlite3.connect('main.db')
    result = pd.read_sql_query(query, conn).to_string()
    conn.close()
    return result 

def generate_pipeline(sql_llm, reviewer_llm, user_input, database_name, table_description, table_content, html_file=None):
    print("Generating SQL query...")
    query = generate_SQL_query(table_description, table_content, user_input, sql_llm)
    print("Validating SQL query...")
    print(user_input, query)
    valid_query = validate_SQL_query(table_description, table_content, user_input, query, reviewer_llm)
    print("Now Executing The Query")
    answer = execute_SQL_query(database_name,valid_query)
    result = str(answer)
    if html_file:
        print("Generating visualizations")
        create_visualization(database_name, user_input, valid_query, reviewer_llm, html_file)
    final_response = generate_response(reviewer_llm, user_input, result)
    print(final_response)
    source_attribution(user_input,valid_query,reviewer_llm,final_response)
    return final_response


def talk_to_my_data(sql_llm, reviewer_llm, table_description, table_content, database_name, user_input):
    def process_question(question, c):
        html_file = f'chart_{c}.html'
        question = question.strip()
        print("Question:", question)
        response = generate_pipeline(sql_llm, reviewer_llm, question, database_name, table_description, table_content, html_file)
        return response, c

    print("Generating sample questions")
    sample_questions = generate_questions(reviewer_llm, table_description, table_content)
    questions = re.findall(r"Question:\s*(.*?)(?=\d+\.\sQuestion:|\Z)", sample_questions, re.DOTALL)
    print(sample_questions)
    with concurrent.futures.ThreadPoolExecutor() as executor:
        futures = []
        c = 1
        for question in questions: 
            if(question!=''):
                futures.append(executor.submit(process_question, question.strip(), c))
                c += 1
        for future in concurrent.futures.as_completed(futures):
            try:
                response, index = future.result()
                with open(f'response_{index}.txt', 'w') as file:
                    file.write(response)
            except Exception as exc:
                print(f'Generated an exception: {exc}')
    html_file = 'src/visuals.html'
    return generate_pipeline(sql_llm, reviewer_llm, user_input, database_name, table_description, table_content, html_file)

def source_attribution(question,query,reviewer_llm,nlp_ans):
    prompt='''When provided with an SQL query, extract and return the names of the tables and the columns used from each table. Present the information clearly in sentence format, without any additional symbols or hashtags.'''
    global tables_used
    tables_used=reviewer_llm.invoke(prompt+query)
    file_path = 'src/metadata.txt'
    with open(file_path, 'w') as file:
        file.write("Question:\n")
        file.write(question + '\n\n')
        file.write("SQL Query:\n")
        file.write(query + '\n\n')
        file.write("Table_Info:")
        file.write(tables_used + '\n')
        file.write("\nResponse:\n")
        file.write(nlp_ans)
    print(f"Content successfully written to {file_path}")



def process_xlsx_file(file):
    if file.filename.endswith('.xlsx'):
        wb = openpyxl.load_workbook(file.stream)
        sheet_names = wb.sheetnames
        return {'fileType': 'XLSX', 'sheetNames': sheet_names}
    else:
        return None

app = Flask(__name__)
CORS(app) 

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part in the request'}), 400
    data_file_path = 'Final_Talk_To_Mydata.xlsx'
    file = request.files['file']
    file.save(data_file_path)
    
    if file.filename == '':
        return jsonify({'error': 'No file selected for uploading'}), 400
    print("Creating database from input files...")
    global database_name
    create_SQL_database(data_file_path, database_name)
    print("Extracting metadata...")
    global table_description
    table_description = get_table_description(desc_file_path)
    global table_content
    table_content = get_table_content(data_file_path)
    print("Creating agents...")
    global sql_llm, reviewer_llm
    sql_llm, reviewer_llm = create_agents()
    processed_data = process_xlsx_file(file)
    if processed_data:
        return jsonify(processed_data), 200
    

@app.route('/download', methods=['GET'])
def download():
    file_path = 'src/metadata.txt'
    if os.path.exists("metadata.txt"):
        os.remove("metadata.txt")
    return send_file(file_path, as_attachment=True, download_name='metadata.txt')

@app.route('/generated-html')
def generated_html():
    with open ("src/visuals.html", encoding='utf-8') as file:
        ans = file.read()
    return Response(ans, content_type='text/html')

@app.route('/get-source')
def get_source():
    return jsonify({'message': tables_used.lstrip()}), 200

@app.route('/echo', methods=['POST'])
def echo_message():
    response_message = ''
    data = request.get_json()
    message = data['message']
    if message.strip() == 'Image':
        response_message = './QR.png'
    elif message!='':
        response_message = talk_to_my_data(sql_llm, reviewer_llm, table_description, table_content, database_name, message)
    print("yay")
    return jsonify({'message': response_message}), 200

if __name__ == '__main__':
    app.run(debug=True)